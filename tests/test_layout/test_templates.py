"""Tests for built-in layout templates."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from excelreport.core.workbook import WorkbookManager
from excelreport.layout.engine import LayoutEngine
from excelreport.layout.templates import (
    layout_dashboard,
    layout_report,
    layout_simple,
    layout_summary_detail,
)
from excelreport.theme.presets import (
    THEME_BUSINESS_BLUE,
    THEME_FRESH_GREEN,
    THEME_MINIMAL_GRAY,
)


class TestLayoutSimple:
    """Tests for layout_simple template."""

    def test_creates_single_sheet(self, temp_xlsx_path: Path) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        df = pd.DataFrame({"A": [1, 2, 3]})
        layout_simple(engine, "Simple Report", df)

        wm = WorkbookManager(temp_xlsx_path)
        engine.build(wm)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        assert "Sheet1" in wb.sheetnames

    def test_works_with_different_themes(self, temp_xlsx_path: Path) -> None:
        for theme in [THEME_BUSINESS_BLUE, THEME_FRESH_GREEN, THEME_MINIMAL_GRAY]:
            engine = LayoutEngine(theme)
            df = pd.DataFrame({"Data": [10, 20]})
            layout_simple(engine, f"Report — {theme.name}", df)

            wm = WorkbookManager(temp_xlsx_path)
            engine.build(wm)
            wm.close()

            assert temp_xlsx_path.stat().st_size > 0


class TestLayoutSummaryDetail:
    """Tests for layout_summary_detail template."""

    def test_creates_two_sheets(self, temp_xlsx_path: Path) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        summary = pd.DataFrame({"Metric": ["Revenue", "Cost"], "Value": [1000, 600]})
        chart = pd.DataFrame({"Month": ["Jan", "Feb", "Mar"], "Sales": [100, 200, 150]})
        detail = pd.DataFrame({"Product": ["A", "B", "C"], "Qty": [5, 8, 3]})

        layout_summary_detail(
            engine,
            "Monthly Report",
            summary,
            chart,
            detail,
            chart_type="column",
            chart_title="Sales Trend",
        )

        wm = WorkbookManager(temp_xlsx_path)
        engine.build(wm)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        assert "Overview" in wb.sheetnames
        assert "Details" in wb.sheetnames


class TestLayoutDashboard:
    """Tests for layout_dashboard template."""

    def test_dashboard_layout(self, temp_xlsx_path: Path) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        kpi = pd.DataFrame({"KPI": ["Sales", "Margin"], "Value": [5000, 0.35]})
        chart1 = pd.DataFrame({"Region": ["N", "S"], "Revenue": [800, 600]})
        chart2 = pd.DataFrame({"Region": ["N", "S"], "Cost": [400, 300]})

        layout_dashboard(
            engine,
            "Dashboard",
            kpi,
            [chart1, chart2],
            chart_types=["column", "bar"],
            chart_titles=["Revenue", "Cost"],
        )

        wm = WorkbookManager(temp_xlsx_path)
        engine.build(wm)
        wm.close()

        assert temp_xlsx_path.stat().st_size > 0


class TestLayoutReport:
    """Tests for layout_report template."""

    def test_report_layout(self, temp_xlsx_path: Path) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        chart = pd.DataFrame({"Month": ["Jan", "Feb"], "Revenue": [5000, 7000]})
        table = pd.DataFrame({"Item": ["Widget", "Gadget"], "Sold": [120, 80]})

        layout_report(
            engine,
            "Q1 2026 Report",
            "Prepared by Analytics Team",
            "This report summarizes the key business metrics for Q1 2026.",
            chart,
            table,
            footer_text="Confidential — Internal Use Only",
            chart_type="line",
            chart_title="Revenue Trend",
        )

        wm = WorkbookManager(temp_xlsx_path)
        engine.build(wm)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        assert "Report" in wb.sheetnames

        ws_r = wb["Report"]
        # Verify title, subtitle
        assert ws_r["B3"].value == "Q1 2026 Report"
        assert ws_r["B6"].value == "Prepared by Analytics Team"
