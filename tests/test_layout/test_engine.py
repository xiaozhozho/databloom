"""Tests for LayoutEngine and SheetLayout."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from excelreport.core.workbook import WorkbookManager
from excelreport.elements.table import TableElement
from excelreport.elements.text import SubtitleElement, TitleElement
from excelreport.layout.engine import LayoutEngine, SheetLayout
from excelreport.theme.presets import THEME_BUSINESS_BLUE


class TestSheetLayout:
    """Tests for SheetLayout."""

    def test_creation(self) -> None:
        sl = SheetLayout("Sheet1", THEME_BUSINESS_BLUE)
        assert sl.sheet_name == "Sheet1"
        assert sl.element_count == 0

    def test_place_row_sequential(self) -> None:
        sl = SheetLayout("S", THEME_BUSINESS_BLUE)
        title = TitleElement("Hello")
        table = TableElement(pd.DataFrame({"A": [1]}))

        p1 = sl.place_row(title)
        p2 = sl.place_row(table)

        assert p1.logical_row == 0
        assert p2.logical_row == 1
        assert sl.element_count == 2

    def test_place_grid_fixed_position(self) -> None:
        sl = SheetLayout("S", THEME_BUSINESS_BLUE)
        el = TitleElement("Fixed")
        placed = sl.place_grid(el, row=3, col=0)
        assert placed.logical_row == 3
        assert sl._next_row == 4  # moved past

    def test_place_grid_does_not_move_next_row_backward(self) -> None:
        sl = SheetLayout("S", THEME_BUSINESS_BLUE)
        sl.place_grid(TitleElement("T"), row=5, col=0)
        sl.place_grid(TitleElement("T2"), row=2, col=0)
        assert sl._next_row == 6  # stays at max

    def test_place_row_full_width_elements_start_at_col_0(self) -> None:
        sl = SheetLayout("S", THEME_BUSINESS_BLUE)
        placed = sl.place_row(TitleElement("X"), col_offset=5)
        # full-width elements ignore col_offset
        assert placed.logical_col == 0

    def test_render_writes_to_sheet(self, temp_xlsx_path: Path) -> None:
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Data")

        sl = SheetLayout("Data", THEME_BUSINESS_BLUE)
        sl.place_row(TitleElement("Report"))
        sl.place_row(TableElement(pd.DataFrame({"X": [10, 20]})))
        sl.render(wm, ws)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Data"]
        # margin_top=2, margin_left=1 → title starts at row=2, col=1 (0-indexed)
        # which is Excel row 3, column B
        assert ws_r["B3"].value == "Report"
        assert ws_r["B6"].value == "X"  # spacing=2, 1 row + 2 spacing → row 5


class TestLayoutEngine:
    """Tests for LayoutEngine."""

    def test_add_sheet_switches_current(self) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        sl = engine.add_sheet("Overview")
        assert engine.current_sheet() is sl

    def test_add_duplicate_sheet_raises(self) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        engine.add_sheet("Data")
        with pytest.raises(ValueError, match="already exists"):
            engine.add_sheet("Data")

    def test_current_sheet_raises_when_empty(self) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        with pytest.raises(RuntimeError, match="No sheet added"):
            engine.current_sheet()

    def test_place_row_adds_to_current_sheet(self) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        engine.add_sheet("S1")
        engine.place_row(TitleElement("T"))
        assert engine.current_sheet().element_count == 1

    def test_build_creates_multiple_sheets(self, temp_xlsx_path: Path) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)

        engine.add_sheet("Overview")
        engine.place_row(TitleElement("Overview Sheet"))
        engine.place_row(TableElement(pd.DataFrame({"A": [1, 2]})))

        engine.add_sheet("Details")
        engine.place_row(TitleElement("Details Sheet"))
        engine.place_row(TableElement(pd.DataFrame({"B": [3, 4, 5]})))

        wm = WorkbookManager(temp_xlsx_path)
        engine.build(wm)
        wm.close()

        assert temp_xlsx_path.exists()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        assert "Overview" in wb.sheetnames
        assert "Details" in wb.sheetnames

    def test_place_fixed_position(self) -> None:
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        engine.add_sheet("Grid")
        result = engine.place(TitleElement("F"), row=2, col=0)
        assert result.logical_row == 2

    def test_build_integration_multiple_elements(self, temp_xlsx_path: Path) -> None:
        """Integration: LayoutEngine builds a multi-element report."""
        engine = LayoutEngine(THEME_BUSINESS_BLUE)
        engine.add_sheet("Main")

        df = pd.DataFrame({"Product": ["A", "B", "C"], "Sales": [100, 200, 150]})
        engine.place_row(TitleElement("Sales Report"))
        engine.place_row(SubtitleElement("Q2 2026"))
        engine.place_row(TableElement(df))

        wm = WorkbookManager(temp_xlsx_path)
        engine.build(wm)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Main"]
        # margin_top=2, margin_left=1, spacing=2
        # Title at row 0: start_row=2 → Excel cell B3
        # Subtitle at row 1: start_row=5 → Excel cell B6
        # Table at row 2: start_row=8 → Excel cell B9
        assert ws_r["B3"].value == "Sales Report"
        assert ws_r["B6"].value == "Q2 2026"
        assert ws_r["B9"].value == "Product"
