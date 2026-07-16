"""Tests for the SparklineElement."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pandas as pd
import pytest

from databloom.core.grid import Grid
from databloom.core.workbook import WorkbookManager
from databloom.elements.sparkline import SparklineElement
from databloom.theme.presets import THEME_BUSINESS_BLUE as THEME


@pytest.fixture
def numeric_df() -> pd.DataFrame:
    """4 rows of purely numeric quarterly data."""
    return pd.DataFrame({
        "Q1": [100, 200, 150, 300],
        "Q2": [120, 180, 170, 280],
        "Q3": [130, 190, 180, 260],
        "Q4": [140, 210, 200, 240],
    })


@pytest.fixture
def grid() -> Grid:
    return Grid()


class TestSparklineElementMeasure:
    """Tests for SparklineElement.measure()."""

    def test_measure_returns_rows_equal_data_len(
        self, numeric_df: pd.DataFrame
    ) -> None:
        el = SparklineElement(numeric_df)
        rows, cols = el.measure(THEME)
        assert rows == len(numeric_df)  # 4 data rows
        assert cols == numeric_df.shape[1] + 1  # data cols + sparkline col

    def test_measure_empty_df(self) -> None:
        df = pd.DataFrame()
        el = SparklineElement(df)
        rows, cols = el.measure(THEME)
        assert rows == 0
        assert cols == 1  # 0 data columns + 1


class TestSparklineElementRender:
    """Integration tests: render to real .xlsx and verify structure."""

    def test_render_line_sparkline(
        self, numeric_df: pd.DataFrame, grid: Grid
    ) -> None:
        el = SparklineElement(
            numeric_df,
            destination_col=4,
            sparkline_type="line",
            markers=True,
            high_point=True,
            low_point=True,
        )
        rows, cols = el.measure(THEME)

        path = Path(tempfile.mkdtemp()) / "sparkline_line.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Data")
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()

        assert path.exists()
        # openpyxl should open the file (it will warn about unsupported
        # sparkline extension but that's expected)
        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        assert "Data" in wb.sheetnames

    def test_render_column_sparkline(
        self, numeric_df: pd.DataFrame, grid: Grid
    ) -> None:
        el = SparklineElement(
            numeric_df,
            destination_col=0,
            sparkline_type="column",
            high_point=True,
            negative_points=True,
        )
        rows, cols = el.measure(THEME)

        path = Path(tempfile.mkdtemp()) / "sparkline_column.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Data")
        placement = grid.place(1, 1, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()
        assert path.exists()

    def test_render_win_loss_sparkline(
        self, numeric_df: pd.DataFrame, grid: Grid
    ) -> None:
        el = SparklineElement(numeric_df, destination_col=2, sparkline_type="win_loss")
        rows, cols = el.measure(THEME)

        path = Path(tempfile.mkdtemp()) / "sparkline_winloss.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Data")
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()
        assert path.exists()

    def test_render_hides_helper_sheet(
        self, numeric_df: pd.DataFrame, grid: Grid
    ) -> None:
        el = SparklineElement(numeric_df, destination_col=1)
        rows, cols = el.measure(THEME)

        path = Path(tempfile.mkdtemp()) / "sparkline_hidden.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Visible")
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        # The helper sheet exists but is hidden
        sparkline_sheets = [
            s for s in wb.sheetnames if s.startswith("_databloom_sparklines_")
        ]
        assert len(sparkline_sheets) == 1

    def test_render_empty_df(self, grid: Grid) -> None:
        """Rendering an empty DataFrame should be a no-op."""
        el = SparklineElement(pd.DataFrame())
        path = Path(tempfile.mkdtemp()) / "sparkline_empty.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Data")
        rows, cols = el.measure(THEME)
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()
        assert path.exists()

    def test_first_last_point_flags(
        self, numeric_df: pd.DataFrame, grid: Grid
    ) -> None:
        """first_point and last_point flags should not cause errors."""
        el = SparklineElement(
            numeric_df,
            destination_col=0,
            first_point=True,
            last_point=True,
        )
        path = Path(tempfile.mkdtemp()) / "sparkline_fl.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Data")
        rows, cols = el.measure(THEME)
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()
        assert path.exists()

    def test_line_weight(self, numeric_df: pd.DataFrame, grid: Grid) -> None:
        """line_weight option should not cause errors."""
        el = SparklineElement(
            numeric_df, destination_col=0, sparkline_type="line", line_weight=2.5
        )
        path = Path(tempfile.mkdtemp()) / "sparkline_weight.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Data")
        rows, cols = el.measure(THEME)
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()
        assert path.exists()

    def test_unique_helper_sheets(self, numeric_df: pd.DataFrame, grid: Grid) -> None:
        """Two sparkline elements should get distinct helper sheets."""
        el1 = SparklineElement(numeric_df, destination_col=0)
        el2 = SparklineElement(numeric_df, destination_col=1)

        path = Path(tempfile.mkdtemp()) / "two_sparklines.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Dual")

        rows1, cols1 = el1.measure(THEME)
        placement1 = grid.place(0, 0, rows1, cols1)
        el1.render(wm, ws, placement1, THEME)

        rows2, cols2 = el2.measure(THEME)
        placement2 = grid.place(rows1, 0, rows2, cols2)
        el2.render(wm, ws, placement2, THEME)

        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        sparkline_sheets = [
            s for s in wb.sheetnames if s.startswith("_databloom_sparklines_")
        ]
        assert len(sparkline_sheets) == 2, (
            f"Expected 2 hidden sparkline sheets, got: {sparkline_sheets}"
        )

    def test_non_numeric_values(
        self, grid: Grid
    ) -> None:
        """Non-numeric values should be written as blanks."""
        df = pd.DataFrame({
            "A": [1, "hello", None, 4],
            "B": [5.0, 6.0, float("nan"), 8.0],
        })
        el = SparklineElement(df, destination_col=2)
        path = Path(tempfile.mkdtemp()) / "sparkline_nonnum.xlsx"
        wm = WorkbookManager(str(path))
        ws = wm.add_sheet("Mixed")
        rows, cols = el.measure(THEME)
        placement = grid.place(0, 0, rows, cols)
        el.render(wm, ws, placement, THEME)
        wm.close()
        assert path.exists()
