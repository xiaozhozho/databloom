"""Tests for text elements: TitleElement, SubtitleElement, ParagraphElement."""

from __future__ import annotations

from pathlib import Path

from excelreport.core.grid import Grid
from excelreport.core.workbook import WorkbookManager
from excelreport.elements.text import ParagraphElement, SubtitleElement, TitleElement
from excelreport.theme.presets import THEME_BUSINESS_BLUE, THEME_FRESH_GREEN


class TestTitleElement:
    """Tests for TitleElement."""

    def test_measure_one_row(self) -> None:
        el = TitleElement("Sales Report")
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows == 1
        assert cols >= 1

    def test_needs_full_width(self) -> None:
        el = TitleElement("Title")
        assert el.needs_full_width() is True

    def test_render_writes_title(self, temp_xlsx_path: Path) -> None:
        """Integration: verify title renders in real xlsx file."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Sheet1")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        placement = grid.place(0, 0, rows=1, cols=10)
        el = TitleElement("Q3 Sales Report")
        el.render(wm, ws, placement, THEME_BUSINESS_BLUE)
        wm.close()

        # Read back with openpyxl
        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Sheet1"]
        assert ws_r["A1"].value == "Q3 Sales Report"
        assert ws_r["A1"].font.bold is True
        assert ws_r["A1"].font.size == 20

    def test_render_multiple_without_crash(self, temp_xlsx_path: Path) -> None:
        """Verify multiple elements render sequentially without error."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Data")
        grid = Grid(margin_top=0, margin_left=0, spacing=1)

        t1 = TitleElement("Main Title")
        t2 = TitleElement("Section Title")

        p1 = grid.place(0, 0, rows=1, cols=10)
        p2 = grid.place(1, 0, rows=1, cols=10)

        t1.render(wm, ws, p1, THEME_BUSINESS_BLUE)
        t2.render(wm, ws, p2, THEME_FRESH_GREEN)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Data"]
        assert ws_r["A1"].value == "Main Title"
        assert ws_r["A3"].value == "Section Title"  # row 0 + spacing=1 => row 3


class TestSubtitleElement:
    """Tests for SubtitleElement."""

    def test_measure_one_row(self) -> None:
        el = SubtitleElement("Overview of Q3 results")
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows == 1

    def test_needs_full_width(self) -> None:
        el = SubtitleElement("Sub")
        assert el.needs_full_width() is True

    def test_render_is_smaller_than_title(self, temp_xlsx_path: Path) -> None:
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("S")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        t_el = TitleElement("Big")
        s_el = SubtitleElement("Small")

        t_el.render(wm, ws, grid.place(0, 0, rows=1, cols=10), THEME_BUSINESS_BLUE)
        s_el.render(wm, ws, grid.place(1, 0, rows=1, cols=10), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["S"]
        assert ws_r["A1"].font.size == 20
        assert ws_r["A2"].font.size == 13  # subtitle smaller


class TestParagraphElement:
    """Tests for ParagraphElement."""

    def test_measure_short_text(self) -> None:
        el = ParagraphElement("Brief note.")
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows == 1

    def test_measure_long_text_multiple_rows(self) -> None:
        long_text = "word " * 100  # ~500 chars, should be ~7 rows
        el = ParagraphElement(long_text)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows > 1

    def test_needs_full_width(self) -> None:
        assert ParagraphElement("text").needs_full_width() is True

    def test_render_paragraph(self, temp_xlsx_path: Path) -> None:
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("P")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = ParagraphElement("This is a note.")
        el.render(wm, ws, grid.place(0, 0, rows=1, cols=10), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["P"]
        assert ws_r["A1"].value == "This is a note."
        assert ws_r["A1"].font.size == 10
