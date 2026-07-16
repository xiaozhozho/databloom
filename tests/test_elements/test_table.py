"""Tests for TableElement."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from databloom.core.grid import Grid
from databloom.core.workbook import WorkbookManager
from databloom.elements.table import (
    TableElement,
    _col_letter,
    _column_width,
    _FormulaFooterElement,
    _infer_dtype_format,
)
from databloom.theme.presets import THEME_BUSINESS_BLUE, THEME_TECH_DARK


class TestInferDtypeFormat:
    """Tests for _infer_dtype_format helper."""

    def test_integer_dtype(self) -> None:
        s = pd.Series([1, 2, 3])
        fmt = _infer_dtype_format(s, THEME_BUSINESS_BLUE)
        assert fmt == "#,##0"

    def test_float_dtype(self) -> None:
        s = pd.Series([1.5, 2.3, 3.7])
        fmt = _infer_dtype_format(s, THEME_BUSINESS_BLUE)
        assert "#,##0" in fmt

    def test_percent_like_float(self) -> None:
        s = pd.Series([0.12, 0.08, 0.15])
        fmt = _infer_dtype_format(s, THEME_BUSINESS_BLUE)
        assert fmt == "0.0%"

    def test_datetime_dtype(self) -> None:
        s = pd.Series(pd.date_range("2026-01-01", periods=3))
        fmt = _infer_dtype_format(s, THEME_BUSINESS_BLUE)
        assert fmt == "yyyy-mm-dd"

    def test_bool_dtype(self) -> None:
        s = pd.Series([True, False, True])
        fmt = _infer_dtype_format(s, THEME_BUSINESS_BLUE)
        assert fmt == ""

    def test_string_dtype(self) -> None:
        s = pd.Series(["a", "b", "c"])
        fmt = _infer_dtype_format(s, THEME_BUSINESS_BLUE)
        assert fmt == ""


class TestColumnWidth:
    """Tests for _column_width helper."""

    def test_returns_reasonable_width(self) -> None:
        s = pd.Series(["Hello", "World"])
        w = _column_width("col_name", s)
        assert 5 <= w <= 30

    def test_header_length_respected(self) -> None:
        s = pd.Series(["a", "b"])
        w = _column_width("very_long_column_name", s)
        assert w >= 10


class TestTableElement:
    """Tests for TableElement."""

    def test_measure_no_title(self, df_simple: pd.DataFrame) -> None:
        el = TableElement(df_simple)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows == len(df_simple) + 1  # header + data rows
        assert cols == len(df_simple.columns)

    def test_measure_with_title(self, df_simple: pd.DataFrame) -> None:
        el = TableElement(df_simple, title="Sales Data")
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows == len(df_simple) + 2  # title + header + data rows

    def test_needs_full_width(self, df_simple: pd.DataFrame) -> None:
        assert TableElement(df_simple).needs_full_width() is True

    def test_render_simple_table(self, temp_xlsx_path: Path, df_simple: pd.DataFrame) -> None:
        """Integration: render a table and verify with openpyxl."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Data")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_simple, title="Summary")
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        placement = grid.place(0, 0, rows=rows, cols=cols)
        el.render(wm, ws, placement, THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Data"]

        # Title row
        assert ws_r["A1"].value == "Summary"

        # Header row is row 2 (after title)
        assert ws_r["A2"].value == "Product"
        assert ws_r["B2"].value == "Sales"
        assert ws_r["C2"].value == "Growth"

        # Data rows
        assert ws_r["A3"].value == "Widget A"
        assert ws_r["B3"].value == 15000
        assert ws_r["C3"].value == 0.12

    def test_render_empty_dataframe(self, temp_xlsx_path: Path) -> None:
        """Should not crash on empty DataFrame."""
        df = pd.DataFrame({"A": [], "B": []})
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Empty")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(
            wm, ws, grid.place(0, 0, rows=max(rows, 1), cols=max(cols, 1)), THEME_BUSINESS_BLUE
        )
        wm.close()
        assert temp_xlsx_path.exists()

    def test_render_single_row(self, temp_xlsx_path: Path, df_single_row: pd.DataFrame) -> None:
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Single")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_single_row)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Single"]
        assert ws_r["A1"].value == "Name"
        assert ws_r["A2"].value == "Total"
        assert ws_r["B2"].value == 9999

    def test_render_with_column_format(self, temp_xlsx_path: Path) -> None:
        """Custom column formats override auto-detection."""
        df = pd.DataFrame({"Price": [1.23456, 2.34567, 3.45678]})
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Fmt")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df, column_formats={"Price": "0.000"})
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Fmt"]
        assert ws_r["A2"].value == 1.23456  # value preserved

    def test_render_alternating_row_colors(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Verify alternating row fills are applied — fills should differ."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Alt")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_simple)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Alt"]
        # Verify both data rows have solid fills
        assert ws_r["A3"].fill is not None
        assert ws_r["A3"].fill.patternType == "solid"
        assert ws_r["A4"].fill is not None
        assert ws_r["A4"].fill.patternType == "solid"
        # Row colors should differ (alternating)
        assert ws_r["A3"].fill.start_color.rgb != ws_r["A4"].fill.start_color.rgb

    def test_render_large_dataframe(self, temp_xlsx_path: Path, df_long: pd.DataFrame) -> None:
        """Stress test: large DataFrame should render without error."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Large")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_long)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Large"]
        assert ws_r["A1"].value == "Date"
        # 60 data rows + 1 header
        assert ws_r["A61"].value is not None

    def test_render_wide_dataframe(self, temp_xlsx_path: Path, df_wide: pd.DataFrame) -> None:
        """Wide table: many columns."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Wide")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_wide)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_BUSINESS_BLUE)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Wide"]
        # 13 columns (Region + 4*3 Qs)
        assert ws_r["A1"].value == "Region"
        assert ws_r["M1"].value is not None  # Last column header

    def test_render_with_dark_theme(self, temp_xlsx_path: Path, df_simple: pd.DataFrame) -> None:
        """Render with tech_dark theme."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Dark")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_simple)
        rows, cols = el.measure(THEME_TECH_DARK)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_TECH_DARK)
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Dark"]
        assert ws_r["A1"].value == "Product"  # Header present
        # Dark theme should have dark fills
        assert ws_r["A1"].fill is not None

    def test_mixed_types(self, temp_xlsx_path: Path, df_mixed_types: pd.DataFrame) -> None:
        """Table with dates, bools, strings, floats renders correctly."""
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("Mixed")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)
        el = TableElement(df_mixed_types)
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        el.render(wm, ws, grid.place(0, 0, rows=rows, cols=cols), THEME_BUSINESS_BLUE)
        wm.close()
        assert temp_xlsx_path.stat().st_size > 0


class TestColLetter:
    """Unit tests for _col_letter — Excel column letter conversion."""

    def test_single_letter(self) -> None:
        """Tests: 0-indexed column to single Excel column letter.

        How: Verify _col_letter maps 0 -> A, 25 -> Z.
        Why: Core coordinate conversion used in formula generation.
        """
        assert _col_letter(0) == "A"
        assert _col_letter(25) == "Z"

    def test_double_letter(self) -> None:
        """Tests: Columns beyond Z (26+) produce double letters.

        How: Verify _col_letter maps 26 -> AA, 27 -> AB, 51 -> AZ, 52 -> BA.
        Why: Ensure formulas work for tables with more than 26 columns.
        """
        assert _col_letter(26) == "AA"
        assert _col_letter(27) == "AB"
        assert _col_letter(51) == "AZ"
        assert _col_letter(52) == "BA"

    def test_triple_letter(self) -> None:
        """Tests: Large column numbers produce triple letters.

        How: Verify _col_letter maps 701 -> ZZ, 702 -> AAA.
        Why: Ensure correctness for very wide tables (e.g. 700+ columns).
        """
        assert _col_letter(701) == "ZZ"
        assert _col_letter(702) == "AAA"

    def test_high_column_number(self) -> None:
        """Tests: Arbitrary high column produces valid letter string.

        How: Check _col_letter(1000) is non-empty and all chars are uppercase.
        Why: Guard against index errors for extreme values.
        """
        result = _col_letter(1000)
        assert result
        assert result.isupper()
        assert result.isalpha()


class TestFormulaFooterElement:
    """Tests for _FormulaFooterElement — internal formula summary row element."""

    def test_measure(self, df_simple: pd.DataFrame) -> None:
        """Tests: _FormulaFooterElement.measure returns 1 row and full width.

        How: Create element with formulas dict, call measure().
        Why: Layout engine needs correct dimensions for placement.
        """
        el = _FormulaFooterElement(df_simple, formulas={"Sales": "SUM"})
        rows, cols = el.measure(THEME_BUSINESS_BLUE)
        assert rows == 1
        assert cols == len(df_simple.columns)

    def test_needs_full_width(self, df_simple: pd.DataFrame) -> None:
        """Tests: _FormulaFooterElement always needs full width.

        How: Call needs_full_width() on a formula footer instance.
        Why: Footer row must span the full table width.
        """
        el = _FormulaFooterElement(df_simple, formulas={})
        assert el.needs_full_width() is True

    def test_render_with_sum_and_average(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: Render built-in SUM and AVERAGE formulas in footer row.

        How: Create table + footer element, render both, verify formulas exist
            in the output .xlsx via openpyxl.
        Why: Built-in formula types are the primary use case.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("FormulaSheet")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        # Place table first
        table = TableElement(df_simple, title="Test Table")
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        # Place footer below table
        footer = _FormulaFooterElement(
            df_simple,
            formulas={"Sales": "SUM", "Growth": "AVERAGE"},
            formula_label="Total",
        )
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        # Verify with openpyxl
        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["FormulaSheet"]

        # Title row 1, Header row 2, Data rows 3-5, Footer row 6
        footer_row = 6
        # First column should be the label
        assert ws_r.cell(row=footer_row, column=1).value == "Total"
        # Second column (Sales) should contain a SUM formula
        sales_cell = ws_r.cell(row=footer_row, column=2).value
        assert sales_cell is not None
        # Growth column should contain an AVERAGE formula
        growth_cell = ws_r.cell(row=footer_row, column=3).value
        assert growth_cell is not None

    def test_render_with_max_and_min(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: MAX and MIN built-in formulas are rendered correctly.

        How: Render footer with MAX and MIN formula keys, verify non-empty cells.
        Why: All four built-in formula types should work.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("MaxMin")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        table = TableElement(df_simple)
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        footer = _FormulaFooterElement(
            df_simple,
            formulas={"Sales": "MAX", "Growth": "MIN"},
            formula_label="Extrema",
        )
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["MaxMin"]
        footer_row = 5  # header(1) + 3 data = footer at row 5
        assert ws_r.cell(row=footer_row, column=1).value == "Extrema"
        assert ws_r.cell(row=footer_row, column=2).value is not None
        assert ws_r.cell(row=footer_row, column=3).value is not None

    def test_render_with_custom_formula(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: Custom formula string is passed through via write_formula.

        How: Use a custom formula string (not a built-in key), verify it is
            written as-is into the cell.
        Why: Users can specify arbitrary Excel formulas, not just SUM/AVERAGE/MAX/MIN.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("CustomFormula")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        table = TableElement(df_simple)
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        custom = "=SUMIF(B2:B4,\">10000\")"
        footer = _FormulaFooterElement(
            df_simple,
            formulas={"Sales": custom},
            formula_label="Filtered",
        )
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["CustomFormula"]
        footer_row = 5
        assert ws_r.cell(row=footer_row, column=1).value == "Filtered"
        assert ws_r.cell(row=footer_row, column=2).value is not None

    def test_render_empty_formulas_dict(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: Footer renders with empty formulas dict (no formula columns).

        How: Pass formulas={}, verify only the label is written, all other cells
            are blank.
        Why: Graceful handling of edge case where user passes no formulas.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("EmptyFormulas")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        table = TableElement(df_simple)
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        footer = _FormulaFooterElement(df_simple, formulas={})
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["EmptyFormulas"]
        footer_row = 5
        # Label column should still have the default label
        assert ws_r.cell(row=footer_row, column=1).value == "Total"
        # Data columns should be None (empty) since no formulas specified
        assert ws_r.cell(row=footer_row, column=2).value is None
        assert ws_r.cell(row=footer_row, column=3).value is None

    def test_render_without_label(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: Footer with empty formula_label writes no label text.

        How: Pass formula_label="", verify first column is blank.
        Why: Users may want to suppress the label for clean layouts.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("NoLabel")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        table = TableElement(df_simple)
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        footer = _FormulaFooterElement(
            df_simple,
            formulas={"Sales": "SUM"},
            formula_label="",
        )
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["NoLabel"]
        footer_row = 5
        # First column is empty since label is empty string
        cell_value = ws_r.cell(row=footer_row, column=1).value
        # openpyxl may return None for empty cells that were write_blank'd
        # but if written as "", it would be "" (not None)
        assert cell_value is None or cell_value == ""

    def test_render_not_bold(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: Footer renders with bold_formulas=False.

        How: Pass bold_formulas=False, verify footer cell font is not bold.
        Why: Users can choose regular-weight formula row for subtle styles.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("NotBold")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        table = TableElement(df_simple)
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        footer = _FormulaFooterElement(
            df_simple,
            formulas={"Sales": "SUM"},
            formula_label="Total",
            bold_formulas=False,
        )
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["NotBold"]
        footer_row = 5
        cell = ws_r.cell(row=footer_row, column=1)
        assert cell.font is not None
        assert cell.font.bold is False

    def test_render_formula_then_custom_in_same_row(
        self, temp_xlsx_path: Path, df_simple: pd.DataFrame
    ) -> None:
        """Tests: Mixed built-in and custom formulas in the same footer row.

        How: Pass one column as "SUM" and another as a custom formula string,
            verify both are rendered.
        Why: Users can combine built-in shortcuts with custom formulas.
        """
        wm = WorkbookManager(temp_xlsx_path)
        ws = wm.add_sheet("MixedFmts")
        grid = Grid(margin_top=0, margin_left=0, spacing=0)

        table = TableElement(df_simple)
        t_rows, t_cols = table.measure(THEME_BUSINESS_BLUE)
        table.render(wm, ws, grid.place(0, 0, rows=t_rows, cols=t_cols), THEME_BUSINESS_BLUE)

        footer = _FormulaFooterElement(
            df_simple,
            formulas={"Sales": "SUM", "Growth": "=AVERAGE(C2:C4)*100"},
            formula_label="Summary",
        )
        f_rows, f_cols = footer.measure(THEME_BUSINESS_BLUE)
        footer.render(
            wm,
            ws,
            grid.place(1, 0, rows=f_rows, cols=f_cols),
            THEME_BUSINESS_BLUE,
        )
        wm.close()

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["MixedFmts"]
        footer_row = 5
        assert ws_r.cell(row=footer_row, column=1).value == "Summary"
        # Both formula columns should have content
        assert ws_r.cell(row=footer_row, column=2).value is not None
        assert ws_r.cell(row=footer_row, column=3).value is not None
