"""Tests for Report builder (declarative API)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from databloom import Report
from databloom.theme.presets import THEME_BUSINESS_BLUE, THEME_TECH_DARK


class TestReportBuilder:
    """Tests for the Report builder class."""

    def test_create_defaults(self) -> None:
        r = Report()
        assert r.title == "Report"
        assert r.theme == THEME_BUSINESS_BLUE

    def test_create_with_theme_string(self) -> None:
        r = Report(theme="tech_dark")
        assert r.theme == THEME_TECH_DARK

    def test_create_with_theme_object(self) -> None:
        r = Report(theme=THEME_TECH_DARK)
        assert r.theme == THEME_TECH_DARK

    def test_create_with_invalid_theme(self) -> None:
        with pytest.raises(KeyError):
            Report(theme="nonexistent")

    def test_add_sheet_returns_self(self) -> None:
        r = Report()
        result = r.add_sheet("Data")
        assert result is r

    def test_add_title_returns_self(self) -> None:
        r = Report().add_sheet("S")
        assert r.add_title("Hello") is r

    def test_add_table_returns_self(self) -> None:
        r = Report().add_sheet("S")
        df = pd.DataFrame({"A": [1]})
        assert r.add_table(df) is r

    def test_add_chart_returns_self(self) -> None:
        r = Report().add_sheet("S")
        df = pd.DataFrame({"X": [1, 2], "Y": [3, 4]})
        assert r.add_chart(df) is r

    def test_add_subtitle_returns_self(self) -> None:
        r = Report().add_sheet("S")
        assert r.add_subtitle("Sub") is r

    def test_add_paragraph_returns_self(self) -> None:
        r = Report().add_sheet("S")
        assert r.add_paragraph("text") is r

    def test_build_returns_bytes_when_no_path(self) -> None:
        r = Report(theme="business_blue")
        r.add_sheet("Data")
        r.add_title("Test")
        r.add_table(pd.DataFrame({"A": [1, 2, 3]}))
        result = r.build()
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_build_writes_file_when_path_given(self, temp_xlsx_path: Path) -> None:
        r = Report(theme="business_blue")
        r.add_sheet("Data")
        r.add_title("Report")
        r.add_table(pd.DataFrame({"X": [10, 20]}))
        result = r.build(temp_xlsx_path)
        assert result is None
        assert temp_xlsx_path.exists()
        assert temp_xlsx_path.stat().st_size > 0

    def test_build_multi_sheet(self, temp_xlsx_path: Path) -> None:
        r = Report(title="Multi-Sheet Report", theme="fresh_green")
        r.add_sheet("Sheet1")
        r.add_title("First Sheet")
        r.add_table(pd.DataFrame({"A": [1]}))

        r.add_sheet("Sheet2")
        r.add_title("Second Sheet")
        r.add_table(pd.DataFrame({"B": [2, 3]}))

        r.build(temp_xlsx_path)

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames

    def test_build_with_chart(self, temp_xlsx_path: Path) -> None:
        r = Report(theme="business_blue")
        r.add_sheet("Chart")
        r.add_title("Sales")
        r.add_chart(
            pd.DataFrame({"Month": ["Jan", "Feb", "Mar"], "Revenue": [100, 200, 150]}),
            type="column",
            category_col="Month",
            value_cols=["Revenue"],
            title="Revenue by Month",
        )
        r.build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_build_full_report(self, temp_xlsx_path: Path) -> None:
        """Integration: build a full multi-element report."""
        r = Report(title="Q3 Analysis", theme="business_blue")
        r.add_sheet("Analysis")
        r.add_title("Q3 2026 Sales Analysis")
        r.add_subtitle("Executive Summary")
        r.add_paragraph(
            "This report provides an overview of Q3 2026 sales performance across all regions."
        )
        r.add_table(
            pd.DataFrame(
                {
                    "Region": ["North", "South", "East", "West"],
                    "Revenue": [125000, 98000, 142000, 87000],
                    "Growth": [0.12, 0.08, 0.15, 0.05],
                }
            ),
            title="Regional Performance",
        )
        r.add_chart(
            pd.DataFrame(
                {
                    "Month": ["Jul", "Aug", "Sep"],
                    "Revenue": [42000, 45000, 48000],
                }
            ),
            type="line",
            category_col="Month",
            value_cols=["Revenue"],
            title="Monthly Revenue Trend",
        )
        r.build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_build_minimal(self) -> None:
        """Build with minimal configuration."""
        r = Report()
        r.add_sheet("Data")
        r.add_table(pd.DataFrame({"X": [1]}))
        data = r.build()
        assert isinstance(data, bytes)

    def test_apply_template_simple(self, temp_xlsx_path: Path) -> None:
        r = Report(theme="business_blue")
        df = pd.DataFrame({"A": [1, 2, 3]})
        r.apply_template("simple", title="Simple", table_df=df)
        r.build(temp_xlsx_path)
        assert temp_xlsx_path.exists()

    def test_apply_template_invalid_raises(self) -> None:
        r = Report()
        with pytest.raises(ValueError, match="not found"):
            r.apply_template("nonexistent")

    def test_add_formula_table_returns_self(self) -> None:
        """Tests: add_formula_table is chainable.

        How: Call add_formula_table on a Report with a DataFrame, verify
            it returns self.
        Why: Builder API requires all add_* methods to return self.
        """
        r = Report().add_sheet("Data")
        df = pd.DataFrame({"Revenue": [100, 200, 300], "Cost": [50, 120, 180]})
        assert r.add_formula_table(df, formulas={"Revenue": "SUM", "Cost": "AVERAGE"}) is r

    def test_add_formula_table_integration(
        self, temp_xlsx_path: Path
    ) -> None:
        """Tests: add_formula_table renders table + formula footer in one step.

        How: Build a report with add_formula_table, verify the output file
            contains both data and formula row via openpyxl.
        Why: This is the primary user-facing API for tables with formula footers.
        """
        df = pd.DataFrame(
            {"Region": ["North", "South", "East"], "Revenue": [1000, 2000, 1500]}
        )
        Report(theme="business_blue").add_sheet("Sheet1").add_formula_table(
            df,
            title="Regional Revenue",
            formulas={"Revenue": "SUM"},
            formula_label="Total",
            bold_formulas=True,
        ).build(temp_xlsx_path)

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Sheet1"]

        # Title in row 3 (margin_top=2 by default), Header in row 4,
        # Data rows 5-7, Footer row 10 (after spacing)
        assert ws_r["B3"].value == "Regional Revenue"
        assert ws_r["B4"].value == "Region"
        assert ws_r["B5"].value == "North"
        # Footer row: label
        assert ws_r.cell(row=10, column=2).value == "Total"
        # Footer row: formula in Revenue column
        assert ws_r.cell(row=10, column=3).value is not None

    def test_add_formula_table_no_formulas(
        self, temp_xlsx_path: Path
    ) -> None:
        """Tests: add_formula_table with no formulas dict still renders.

        How: Call add_formula_table without passing formulas, verify file
            is created successfully.
        Why: formulas parameter defaults to {} so the method should not crash
            when omitted.
        """
        df = pd.DataFrame({"A": [1, 2, 3]})
        Report().add_sheet("Data").add_formula_table(df).build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_add_combo_chart_returns_self(self) -> None:
        """Tests: add_combo_chart is chainable.

        How: Call add_combo_chart with valid bar/line columns, verify
            it returns self.
        Why: Builder API requires all add_* methods to return self.
        """
        r = Report().add_sheet("S")
        df = pd.DataFrame(
            {"Month": ["Jan", "Feb", "Mar"], "Revenue": [100, 200, 150], "Margin%": [0.1, 0.2, 0.15]}
        )
        assert r.add_combo_chart(df, bar_cols=["Revenue"], line_cols=["Margin%"]) is r

    def test_add_combo_chart_integration(
        self, temp_xlsx_path: Path
    ) -> None:
        """Tests: add_combo_chart renders a combo chart into the workbook.

        How: Build a report with add_combo_chart, verify the file is non-empty.
        Why: The combo chart path through the Report builder should be exercised
            to ensure it composes correctly with the layout engine.
        """
        df = pd.DataFrame(
            {
                "Month": ["Jan", "Feb", "Mar", "Apr"],
                "Revenue": [1000, 1500, 1200, 1800],
                "Margin%": [0.12, 0.15, 0.14, 0.18],
            }
        )
        Report(theme="business_blue").add_sheet("Combo").add_title(
            "Performance"
        ).add_combo_chart(
            df,
            category_col="Month",
            bar_cols=["Revenue"],
            line_cols=["Margin%"],
            bar_title="Revenue ($)",
            line_title="Margin",
            title="Revenue vs Margin",
        ).build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_add_image_returns_self(self, tmp_path: Path) -> None:
        """Tests: add_image is chainable.

        How: Create a tiny PNG in a temp dir, call add_image, verify
            it returns self.
        Why: Builder API requires all add_* methods to return self.
        """
        from PIL import Image

        img_path = tmp_path / "test_image.png"
        img = Image.new("RGB", (10, 10), color="blue")
        img.save(img_path)

        r = Report().add_sheet("Images")
        assert r.add_image(str(img_path)) is r

    def test_add_spacer_returns_self(self) -> None:
        """Tests: add_spacer is chainable.

        How: Call add_spacer, verify it returns self.
        Why: Builder API requires all add_* methods to return self.
        """
        r = Report().add_sheet("Sheet")
        assert r.add_spacer(rows=2, height=15) is r

    def test_add_spacer_integration(self, temp_xlsx_path: Path) -> None:
        """Tests: add_spacer adds vertical gap without errors.

        How: Insert a spacer between two elements, build the report,
            verify file is created.
        Why: Spacer is a commonly used layout element that needs integration
            coverage.
        """
        Report().add_sheet("Sheet").add_title("Top").add_spacer(
            rows=2, height=24
        ).add_paragraph("After spacer").build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_set_page_setup_returns_self(self) -> None:
        """Tests: set_page_setup is chainable.

        How: Call set_page_setup with landscape orientation, verify
            it returns self.
        Why: Builder API requires configuration methods to return self.
        """
        r = Report()
        assert r.set_page_setup(orientation="landscape") is r

    def test_set_page_setup_integration(
        self, temp_xlsx_path: Path
    ) -> None:
        """Tests: Page setup settings are applied to the workbook.

        How: Call set_page_setup with custom margins and orientation, build
            the report, verify via openpyxl that landscape orientation was set.
        Why: Page setup is important for print-ready reports and must be
            tested end-to-end.
        """
        Report(theme="business_blue").add_sheet("Sheet1").add_title(
            "Print Test"
        ).set_page_setup(
            orientation="landscape",
            paper=9,  # A4
            margin_left=0.5,
            margin_right=0.5,
            fit_to_width=1,
            fit_to_height=0,
            print_title_rows=1,
        ).build(temp_xlsx_path)

        import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx_path)
        ws_r = wb["Sheet1"]
        # openpyxl reads page setup
        assert ws_r.page_setup.orientation == "landscape"

    def test_set_page_setup_defaults(self, temp_xlsx_path: Path) -> None:
        """Tests: set_page_setup with no args uses defaults from settings.

        How: Call set_page_setup without any keyword arguments, build,
            verify the file is created without errors.
        Why: All parameters have defaults — ensure the method works with
            zero-config calls.
        """
        Report().add_sheet("S").add_table(
            pd.DataFrame({"A": [1]})
        ).set_page_setup().build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_apply_template_dashboard(self, temp_xlsx_path: Path) -> None:
        """Tests: apply_template with 'dashboard' template.

        How: Call apply_template("dashboard") with required kwargs, build,
            verify the file exists.
        Why: All template types should be tested for basic rendering.
        """
        kpi = pd.DataFrame({"Metric": ["Revenue", "Cost"], "Value": [1000, 600]})
        chart_df = pd.DataFrame({"Month": ["Jan", "Feb"], "Sales": [100, 200]})
        Report(theme="business_blue").apply_template(
            "dashboard",
            title="Dashboard Report",
            kpi_df=kpi,
            chart_dfs=[chart_df],
        ).build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_apply_template_report(self, temp_xlsx_path: Path) -> None:
        """Tests: apply_template with 'report' template.

        How: Call apply_template("report") with required kwargs, build,
            verify the file exists.
        Why: The 'report' template is a key layout template.
        """
        chart_df = pd.DataFrame({"Month": ["Jan", "Feb"], "Sales": [100, 200]})
        table_df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        Report().apply_template(
            "report",
            title="Report Title",
            subtitle="Subtitle Text",
            body_text="This is the body.",
            chart_df=chart_df,
            table_df=table_df,
        ).build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_apply_template_summary_detail(self, temp_xlsx_path: Path) -> None:
        """Tests: apply_template with 'summary_detail' template and kwargs.

        How: Call apply_template("summary_detail") with DataFrame kwargs,
            build and verify file.
        Why: summary_detail is the most feature-rich template and needs
            integration coverage.
        """
        summary = pd.DataFrame({"Metric": ["Revenue", "Cost"], "Current": [1000, 600]})
        detail = pd.DataFrame(
            {"Date": ["Jan", "Feb"], "Revenue": [500, 500], "Cost": [300, 300]}
        )
        Report(theme="business_blue").apply_template(
            "summary_detail",
            title="Summary",
            summary_df=summary,
            detail_df=detail,
            chart_df=detail,
            chart_type="column",
        ).build(temp_xlsx_path)
        assert temp_xlsx_path.stat().st_size > 0

    def test_quick_with_theme_object(self) -> None:
        """Tests: Report.quick accepts a Theme instance directly.

        How: Pass a Theme object (not a string) to Report.quick, verify
            the returned Report uses that theme.
        Why: The theme parameter supports both string names and Theme instances.
        """
        df = pd.DataFrame({"A": [1, 2, 3]})
        report = Report.quick(df, theme=THEME_TECH_DARK)
        assert report.theme == THEME_TECH_DARK

    def test_quick_with_many_dataframes(self) -> None:
        """Tests: Report.quick with 3+ DataFrames emits a warning.

        How: Pass 3 DataFrames to Report.quick, verify a UserWarning is raised
            about only using the first 2.
        Why: The quick builder only supports up to 2 DataFrames and should
            warn about this limitation.
        """
        d1 = pd.DataFrame({"A": [1]})
        d2 = pd.DataFrame({"B": [2]})
        d3 = pd.DataFrame({"C": [3]})
        with pytest.warns(UserWarning, match="first 2"):
            Report.quick(d1, d2, d3)
