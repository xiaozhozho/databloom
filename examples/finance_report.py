#!/usr/bin/env python3
"""财务分析报告 — 利润表 + 财务指标 + 组合图表

数据通过 load_dataset("finance_profit") 和 load_dataset("finance_metrics")
从内置 pickle 数据源加载，保证每次运行数据完全一致。

展示 databloom v0.2 的核心新特性：
  - 组合图表（柱状图+折线图双Y轴）
  - 打印页面设置
  - finance_charcoal 主题
  - 列对齐智能推断（数字右对齐，文字左对齐）

Sheet 1 — 利润表:  月度营收/成本/利润 + 组合图表
Sheet 2 — 财务指标:  偿债/营运/盈利/成长能力指标
Sheet 3 — 分析报告:  财务分析文档
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from databloom import Report
from databloom.data import load_dataset

# ── 加载数据（从 pickle 缓存） ──────────────────────────────────────────
profit_df = load_dataset("finance_profit")
fin_metrics = load_dataset("finance_metrics")

# 年度汇总
annual_df = pd.DataFrame({
    "指标": [
        "营业收入(万元)", "营业成本(万元)", "毛利润(万元)", "毛利率",
        "费用合计(万元)", "营业利润(万元)", "营业利润率",
        "净利润(万元)", "净利率",
    ],
    "数值": [
        round(profit_df["营业收入"].sum() / 10000, 1),
        round(profit_df["营业成本"].sum() / 10000, 1),
        round(profit_df["毛利润"].sum() / 10000, 1),
        f"{profit_df['毛利润'].sum() / profit_df['营业收入'].sum():.1%}",
        round(profit_df["费用合计"].sum() / 10000, 1),
        round(profit_df["营业利润"].sum() / 10000, 1),
        f"{profit_df['营业利润'].sum() / profit_df['营业收入'].sum():.1%}",
        round(profit_df["净利润"].sum() / 10000, 1),
        f"{profit_df['净利润'].sum() / profit_df['营业收入'].sum():.1%}",
    ],
})

# ==============================================================================
# 构建报告
# ==============================================================================

report = Report(title="2025年度财务分析报告", theme="finance_charcoal")
report.set_page_setup(orientation="landscape", fit_to_width=1, fit_to_height=0)

# Sheet 1 — 利润表
report.add_sheet("利润表")
report.add_title("2025年度利润表分析")
report.add_subtitle(
    "数据期间：2025年1月—12月 | 数据来源：load_dataset('finance_profit') | 单位：元"
)
report.add_table(profit_df, title="月度利润表（单位：元）", column_formats={
    "营业收入": "#,##0", "营业成本": "#,##0", "毛利润": "#,##0",
    "毛利率": "0.0%", "销售费用": "#,##0", "管理费用": "#,##0",
    "研发费用": "#,##0", "费用合计": "#,##0", "营业利润": "#,##0",
    "营业利润率": "0.0%", "净利润": "#,##0", "净利率": "0.0%",
})
report.add_spacer(rows=1, height=8)

# 组合图表：柱状图(营收+成本+费用) + 折线图(净利率)
report.add_combo_chart(
    profit_df,
    category_col="月份",
    bar_cols=["营业收入", "营业成本", "费用合计"],
    line_cols=["净利率"],
    bar_title="金额（元）",
    line_title="净利率",
    title="月度经营趋势：营收/成本 vs 净利率（组合图表）",
)

# Sheet 2 — 财务指标
report.add_sheet("财务指标")
report.add_title("2025年度核心财务指标")
report.add_subtitle(
    "偿债能力 · 营运能力 · 盈利能力 · 成长能力 | "
    "数据来源：load_dataset('finance_metrics')"
)
report.add_spacer(rows=1, height=6)
report.add_table(annual_df, title="年度利润表摘要")
report.add_spacer(rows=1, height=8)
report.add_table(fin_metrics, title="财务指标对比分析（2025 vs 2024）")

# Sheet 3 — 分析报告
report.add_sheet("分析报告")
report.add_title("2025年度财务分析报告")
report.add_subtitle("编制日期：2026年1月20日 | 编制部门：财务部")

report.add_paragraph(
    "【一、经营概况】\n\n"
    f"2025年度实现营业收入{profit_df['营业收入'].sum()/10000:,.0f}万元，同比增长18.5%，"
    f"超额完成年度目标(15%增长)。综合毛利率{profit_df['毛利润'].sum()/profit_df['营业收入'].sum():.1%}，"
    f"较上年提升1.4个百分点。全年净利润{profit_df['净利润'].sum()/10000:,.0f}万元，"
    f"净利率{profit_df['净利润'].sum()/profit_df['营业收入'].sum():.1%}，同比增长22.3%。"
    f"ROE达15.8%，较上年提升1.6个百分点，资本使用效率持续改善。"
)

report.add_paragraph(
    "【二、盈利能力分析】\n\n"
    f"全年毛利率呈V型走势，Q1-Q2受原材料涨价影响毛利率阶段性承压"
    f"(低点{profit_df['毛利率'].min():.1%})，下半年供应链优化后稳步回升至"
    f"{profit_df['毛利率'].iloc[-3:].mean():.1%}%。\n\n"
    f"费用率控制良好：销售费用率12%、管理费用率8%、研发费用率3%，三项合计23%，"
    f"处于行业较低水平。研发费用同比增长15%，体现了对产品和技术的持续投入。"
)

report.add_paragraph(
    "【三、风险提示】\n\n"
    "1. 资产负债率42.3%，较上年下降2.8个百分点，财务杠杆处于合理区间；\n"
    "2. Q3营收增速放缓至12%，主要受行业淡季和竞品促销影响；\n"
    "3. 应收账款周转率8.2次，同比提升5.1%，但仍低于行业标杆(10次)；\n"
    "4. 所得税优惠政策2025年底到期，2026年起实际税负率可能上升3-5个百分点。"
)

report.add_paragraph(
    "【四、2026年展望】\n\n"
    f"基于2025年经营数据，预计2026年营收增长10-15%，"
    f"目标{profit_df['营业收入'].sum()*1.12/10000:,.0f}万元。"
    f"核心举措：①优化产品结构，高毛利产品占比从35%提升至42%；"
    f"②深化供应链数字化，采购成本降低3-5%；"
    f"③精细化管理，三项费用率控制在22%以内。"
)

# ==============================================================================
# Build
# ==============================================================================

output_dir = Path(__file__).resolve().parent.parent / "output"
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / "finance_report.xlsx"

report.build(output_path)

size_kb = output_path.stat().st_size / 1024
wb = openpyxl.load_workbook(output_path)
visible = [s for s in wb.sheetnames if not s.startswith("_databloom")]
print(f"✅ 财务分析报告已生成：{output_path}")
print(f"   文件大小：{size_kb:.1f} KB")
print(f"   包含 {len(visible)} 个Sheet：", " | ".join(visible))
for s in visible:
    ws = wb[s]
    print(f"     {s}: {ws.max_row}行 × {ws.max_column}列")
